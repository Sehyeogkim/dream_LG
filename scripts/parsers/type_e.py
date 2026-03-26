"""Type E parser: 벽걸이에어컨.

Row 7+: data. Cols: B=구분, C=구분, D=모델명, E=서비스타입
Price cols (period, qty):
  3yr: F(1), K(2-3), O(4-9), S(10-29), W(30+)
  4yr: AA(1), AF(2-3), AJ(4-9), AN(10-29), AR(30+)
  5yr: AV(1), BA(2-3), BE(4-9), BI(10-29), BM(30+)
  6yr: BQ(1), BV(2-3), BZ(4-9), CD(10-29), CH(30+)
"""
from openpyxl.utils import column_index_from_string as _col

PRICE_COLS = {
    '36': {'1': _col('F'), '2-3': _col('K'), '4-9': _col('O'), '10-29': _col('S'), '30+': _col('W')},
    '48': {'1': _col('AA'), '2-3': _col('AF'), '4-9': _col('AJ'), '10-29': _col('AN'), '30+': _col('AR')},
    '60': {'1': _col('AV'), '2-3': _col('BA'), '4-9': _col('BE'), '10-29': _col('BI'), '30+': _col('BM')},
    '72': {'1': _col('BQ'), '2-3': _col('BV'), '4-9': _col('BZ'), '10-29': _col('CD'), '30+': _col('CH')},
}

DATA_START = 7


def parse_type_e(ws, sheet_name):
    products = []
    current_category = ''
    current_name = ''
    current_model = ''

    for row_idx in range(DATA_START, ws.max_row + 1):
        b_val = ws.cell(row=row_idx, column=_col('B')).value
        c_val = ws.cell(row=row_idx, column=_col('C')).value
        d_val = ws.cell(row=row_idx, column=_col('D')).value
        e_val = ws.cell(row=row_idx, column=_col('E')).value
        f_val = ws.cell(row=row_idx, column=_col('F')).value

        if b_val:
            current_category = str(b_val).strip()
        if c_val:
            current_name = str(c_val).strip().replace('\n', ' ')
        if d_val:
            current_model = str(d_val).strip()

        if f_val is None:
            continue

        service_type = str(e_val).strip() if e_val else ''

        prices = {}
        for period, tier_cols in PRICE_COLS.items():
            period_prices = {}
            for tier, col_idx in tier_cols.items():
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None and isinstance(val, (int, float)):
                    period_prices[tier] = int(val)
            if period_prices:
                prices[period] = period_prices

        if not prices:
            continue

        products.append({
            'model_id': current_model,
            'product_name': f"{current_category} {current_name}".strip(),
            'service_type': service_type,
            'prices': prices,
        })

    return products
