"""제습기 parser.

Row 6+. B=구분, C=모델명, D=방문주기
Price cols:
  3yr: E(1), F(2-3), J(4-9), N(10-29), R(30+)
  4yr: V(1), W(2-3), AA(4-9), AE(10-29), AI(30+)
  5yr: AM(1), AN(2-3), AR(4-9), AV(10-29), AZ(30+)
  6yr: BD(1), BE(2-3), BI(4-9), BM(10-29), BQ(30+)
"""
from openpyxl.utils import column_index_from_string as _col

PRICE_COLS = {
    '36': {'1': _col('E'), '2-3': _col('F'), '4-9': _col('J'), '10-29': _col('N'), '30+': _col('R')},
    '48': {'1': _col('V'), '2-3': _col('W'), '4-9': _col('AA'), '10-29': _col('AE'), '30+': _col('AI')},
    '60': {'1': _col('AM'), '2-3': _col('AN'), '4-9': _col('AR'), '10-29': _col('AV'), '30+': _col('AZ')},
    '72': {'1': _col('BD'), '2-3': _col('BE'), '4-9': _col('BI'), '10-29': _col('BM'), '30+': _col('BQ')},
}

DATA_START = 6


def parse_type_dehumidifier(ws, sheet_name):
    products = []
    current_name = ''
    current_model = ''

    for row_idx in range(DATA_START, ws.max_row + 1):
        b_val = ws.cell(row=row_idx, column=_col('B')).value
        c_val = ws.cell(row=row_idx, column=_col('C')).value
        d_val = ws.cell(row=row_idx, column=_col('D')).value
        e_val = ws.cell(row=row_idx, column=_col('E')).value

        if b_val:
            current_name = str(b_val).strip().replace('\n', ' ')
        if c_val:
            current_model = str(c_val).strip()

        if e_val is None:
            continue

        try:
            vc = int(d_val) if d_val else 0
            visit_cycle = f"{vc}개월" if vc > 0 else "방문없음"
        except (ValueError, TypeError):
            visit_cycle = str(d_val).strip() if d_val else ''

        prices = {}
        for period, tier_cols in PRICE_COLS.items():
            period_prices = {}
            for tier, col_idx in tier_cols.items():
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None and isinstance(val, (int, float)):
                    period_prices[tier] = int(val)
            if period_prices:
                prices[period] = period_prices

        if not prices:
            continue

        products.append({
            'model_id': current_model,
            'product_name': current_name,
            'visit_cycle': visit_cycle,
            'prices': prices,
        })

    return products
