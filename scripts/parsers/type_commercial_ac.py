"""Commercial AC parser: 상업용에어컨.

Structure:
  Row 3-5: headers
  Row 6+: data
  Cols: B=제품군, C=모델명, D=방문주기, E=서비스타입
  Only 3yr contract
  Price cols: G(2~3대)
"""
from openpyxl.utils import column_index_from_string


def _col(letter):
    return column_index_from_string(letter)


DATA_START = 6


def parse_type_commercial_ac(ws, sheet_name):
    products = []
    current_product_group = ''

    for row_idx in range(DATA_START, ws.max_row + 1):
        b_val = ws.cell(row=row_idx, column=_col('B')).value
        c_val = ws.cell(row=row_idx, column=_col('C')).value  # 모델명
        d_val = ws.cell(row=row_idx, column=_col('D')).value  # 방문주기
        e_val = ws.cell(row=row_idx, column=_col('E')).value  # 서비스타입
        f_val = ws.cell(row=row_idx, column=_col('F')).value  # 기본 월요금
        g_val = ws.cell(row=row_idx, column=_col('G')).value  # 2~3대 할인가

        if b_val:
            current_product_group = str(b_val).strip().replace('\n', ' ')

        if f_val is None:
            continue

        model_id = str(c_val).strip() if c_val else ''
        if not model_id:
            continue

        service_type = str(e_val).strip() if e_val else ''
        visit_cycle = str(d_val).strip() if d_val else ''

        prices = {}
        if g_val is not None and isinstance(g_val, (int, float)):
            prices['36'] = int(g_val)

        if not prices:
            continue

        products.append({
            'model_id': model_id,
            'product_name': current_product_group,
            'service_type': service_type,
            'visit_cycle': visit_cycle,
            'prices': prices,
        })

    return _deduplicate(products)


def _deduplicate(products):
    by_key = {}
    for p in products:
        key = (p['model_id'], p['service_type'])
        if key not in by_key:
            by_key[key] = p
    return list(by_key.values())
