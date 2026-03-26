"""Styler/ShuCare parser.

Row 6+: data. D=제품, E=모델명, F=서비스타입, G=방문주기
Price cols:
  3yr: H(1), J(2-3), M(4-9), P(10-29), S(30+)
  4yr: V(1), X(2-3), AA(4-9), AD(10-29), AG(30+)
  5yr: AJ(1), AL(2-3), AO(4-9), AR(10-29), AU(30+)
  6yr: AX(1), AZ(2-3), BC(4-9), BF(10-29), BI(30+)
"""
from openpyxl.utils import column_index_from_string as _col

PRICE_COLS = {
    '36': {'1': _col('H'), '2-3': _col('J'), '4-9': _col('M'), '10-29': _col('P'), '30+': _col('S')},
    '48': {'1': _col('V'), '2-3': _col('X'), '4-9': _col('AA'), '10-29': _col('AD'), '30+': _col('AG')},
    '60': {'1': _col('AJ'), '2-3': _col('AL'), '4-9': _col('AO'), '10-29': _col('AR'), '30+': _col('AU')},
    '72': {'1': _col('AX'), '2-3': _col('AZ'), '4-9': _col('BC'), '10-29': _col('BF'), '30+': _col('BI')},
}

DATA_START = 6


def parse_type_styler(ws, sheet_name):
    products = []
    current_product = ''

    for row_idx in range(DATA_START, ws.max_row + 1):
        d_val = ws.cell(row=row_idx, column=_col('D')).value
        e_val = ws.cell(row=row_idx, column=_col('E')).value
        f_val = ws.cell(row=row_idx, column=_col('F')).value
        g_val = ws.cell(row=row_idx, column=_col('G')).value
        h_val = ws.cell(row=row_idx, column=_col('H')).value

        if d_val:
            current_product = str(d_val).strip().replace('\n', ' ')

        if not e_val or h_val is None:
            continue

        model_id = str(e_val).strip()
        service_type = str(f_val).strip() if f_val else ''
        visit_cycle = str(g_val).strip() if g_val else ''

        prices = {}
        for period, tier_cols in PRICE_COLS.items():
            period_prices = {}
            for tier, col_idx in tier_cols.items():
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None and isinstance(val, (int, float)):
                    period_prices[tier] = int(val)
            if period_prices:
                prices[period] = period_prices

        if not prices:
            continue

        products.append({
            'model_id': model_id,
            'product_name': current_product,
            'service_type': service_type,
            'visit_cycle': visit_cycle,
            'prices': prices,
        })

    return products
