"""Type B parser: 세탁기/건조기.

Same column layout as Type A but E = 서비스타입.
"""
from openpyxl.utils import column_index_from_string as _col

PRICE_COLS = {
    '36': {'1': _col('G'), '2-3': _col('H'), '4-9': _col('L'), '10-29': _col('P'), '30+': _col('T')},
    '48': {'1': _col('X'), '2-3': _col('Y'), '4-9': _col('AC'), '10-29': _col('AG'), '30+': _col('AK')},
    '60': {'1': _col('AO'), '2-3': _col('AP'), '4-9': _col('AT'), '10-29': _col('AX'), '30+': _col('BB')},
    '72': {'1': _col('BF'), '2-3': _col('BG'), '4-9': _col('BK'), '10-29': _col('BO'), '30+': _col('BS')},
}

DATA_START = 7


def parse_type_b(ws, sheet_name):
    products = []
    current_category = ''
    current_name = ''
    current_model = ''

    for row_idx in range(DATA_START, ws.max_row + 1):
        b_val = ws.cell(row=row_idx, column=_col('B')).value
        c_val = ws.cell(row=row_idx, column=_col('C')).value
        d_val = ws.cell(row=row_idx, column=_col('D')).value
        e_val = ws.cell(row=row_idx, column=_col('E')).value
        f_val = ws.cell(row=row_idx, column=_col('F')).value
        g_val = ws.cell(row=row_idx, column=_col('G')).value

        if b_val:
            current_category = str(b_val).strip()
        if c_val:
            current_name = str(c_val).strip().replace('\n', ' ')
        if d_val:
            current_model = str(d_val).strip()

        if g_val is None:
            continue

        service_type = str(e_val).strip() if e_val else ''
        try:
            visit_cycle = int(f_val) if f_val and f_val != 0 else 0
        except (ValueError, TypeError):
            visit_cycle = 0
        visit_str = f"{visit_cycle}개월" if visit_cycle > 0 else "방문없음"

        prices = {}
        for period, tier_cols in PRICE_COLS.items():
            period_prices = {}
            for tier, col_idx in tier_cols.items():
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None and isinstance(val, (int, float)):
                    period_prices[tier] = int(val)
            if period_prices:
                prices[period] = period_prices

        if not prices:
            continue

        products.append({
            'model_id': current_model,
            'product_name': current_name,
            'category_sub': current_category,
            'service_type': service_type,
            'visit_cycle': visit_str,
            'prices': prices,
        })

    return products
