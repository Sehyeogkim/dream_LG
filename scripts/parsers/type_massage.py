"""Massage chair parser.

Row 6+. C=제품군, D=모델명, E=방문주기
Price cols:
  3yr: F(1), H(2-3), K(4-9), N(10-29), Q(30+)
  4yr: T(1), V(2-3), Y(4-9), AB(10-29), AE(30+)
  5yr: AH(1), AJ(2-3), AM(4-9), AP(10-29), AS(30+)
  6yr: AV(1), AX(2-3), BA(4-9), BD(10-29), BG(30+)
"""
from openpyxl.utils import column_index_from_string as _col

PRICE_COLS = {
    '36': {'1': _col('F'), '2-3': _col('H'), '4-9': _col('K'), '10-29': _col('N'), '30+': _col('Q')},
    '48': {'1': _col('T'), '2-3': _col('V'), '4-9': _col('Y'), '10-29': _col('AB'), '30+': _col('AE')},
    '60': {'1': _col('AH'), '2-3': _col('AJ'), '4-9': _col('AM'), '10-29': _col('AP'), '30+': _col('AS')},
    '72': {'1': _col('AV'), '2-3': _col('AX'), '4-9': _col('BA'), '10-29': _col('BD'), '30+': _col('BG')},
}

DATA_START = 6


def parse_type_massage(ws, sheet_name):
    products = []
    current_product_group = ''

    for row_idx in range(DATA_START, ws.max_row + 1):
        c_val = ws.cell(row=row_idx, column=_col('C')).value
        d_val = ws.cell(row=row_idx, column=_col('D')).value
        e_val = ws.cell(row=row_idx, column=_col('E')).value
        f_val = ws.cell(row=row_idx, column=_col('F')).value

        if c_val:
            current_product_group = str(c_val).strip().replace('\n', ' ')

        if not d_val or f_val is None:
            continue

        model_id = str(d_val).strip().split('\n')[0]
        visit_cycle = str(e_val).strip() if e_val else ''

        prices = {}
        for period, tier_cols in PRICE_COLS.items():
            period_prices = {}
            for tier, col_idx in tier_cols.items():
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None and isinstance(val, (int, float)):
                    period_prices[tier] = int(val)
            if period_prices:
                prices[period] = period_prices

        if not prices:
            continue

        products.append({
            'model_id': model_id,
            'product_name': current_product_group,
            'visit_cycle': visit_cycle,
            'prices': prices,
        })

    return products
