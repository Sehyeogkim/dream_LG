"""Type D parser: TV.

Structure:
  Rows 3-4: headers
  Row 5+: data in groups of 5 rows (one per quantity tier)
  Group structure:
    Row 1: B=구분, E=설치유형, F=1대, G=5yr price, J=6yr price
    Row 2: F=2~3대, G=5yr price, J=6yr price
    Row 3: C=본체모델명, D=주문모델명, F=4~9대, ...
    Row 4: F=10~29대, ...
    Row 5: F=30대이상, ...
  We extract the "2~3대" tier and grab model from the adjacent row.
"""
from openpyxl.utils import column_index_from_string


def _col(letter):
    return column_index_from_string(letter)


DATA_START = 5


def parse_type_d(ws, sheet_name):
    products = []

    # First pass: collect all data rows with their values
    rows_data = []
    for row_idx in range(DATA_START, ws.max_row + 1):
        row = {
            'idx': row_idx,
            'B': ws.cell(row=row_idx, column=_col('B')).value,
            'C': ws.cell(row=row_idx, column=_col('C')).value,
            'D': ws.cell(row=row_idx, column=_col('D')).value,
            'E': ws.cell(row=row_idx, column=_col('E')).value,
            'F': ws.cell(row=row_idx, column=_col('F')).value,
            'G': ws.cell(row=row_idx, column=_col('G')).value,
            'J': ws.cell(row=row_idx, column=_col('J')).value,
            'M': ws.cell(row=row_idx, column=_col('M')).value,
        }
        rows_data.append(row)

    # Process in groups: detect groups by B column (product name) appearance
    current_category = ''
    current_install = ''
    i = 0
    while i < len(rows_data):
        row = rows_data[i]

        if row['B']:
            current_category = str(row['B']).strip().replace('\n', ' ')
        if row['E']:
            current_install = str(row['E']).strip().replace('\n', '/')

        f_str = str(row['F']).strip() if row['F'] else ''

        if '2~3대' in f_str or '2~3' in f_str:
            # This is our target row - get prices
            prices = {}
            if row['G'] is not None and isinstance(row['G'], (int, float)):
                prices['60'] = int(row['G'])
            if row['J'] is not None and isinstance(row['J'], (int, float)):
                prices['72'] = int(row['J'])

            # Look for model_id in nearby rows (next 1-2 rows typically)
            model_id = ''
            for offset in range(1, 3):
                if i + offset < len(rows_data):
                    c_val = rows_data[i + offset]['C']
                    d_val = rows_data[i + offset]['D']
                    if c_val:
                        model_id = str(c_val).strip().split('\n')[0]
                        break
                    if d_val:
                        model_id = str(d_val).strip().split('\n')[0]
                        break
            # Also check previous rows
            if not model_id:
                for offset in range(1, 3):
                    if i - offset >= 0:
                        c_val = rows_data[i - offset]['C']
                        d_val = rows_data[i - offset]['D']
                        if c_val:
                            model_id = str(c_val).strip().split('\n')[0]
                            break
                        if d_val:
                            model_id = str(d_val).strip().split('\n')[0]
                            break

            if prices:
                products.append({
                    'model_id': model_id,
                    'product_name': current_category,
                    'install_type': current_install,
                    'prices': prices,
                })

        i += 1

    return products
