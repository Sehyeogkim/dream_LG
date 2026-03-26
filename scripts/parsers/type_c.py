"""Type C parser: 냉장고, 광파오븐, 식기세척기, 전기레인지.

Row 9+: data. Cols: D=구분1, E=구분2, F=모델명, G=서비스타입, H=방문주기

Price columns per (period, qty_tier):
냉장고:   3yr: I(1대),K(2-3),N(4-9),Q(10-29),T(30+) | 4yr: W,Y,AB,AE,AH | 5yr: AK,AM,AP,AS,AV | 6yr: AY,BA,BD,BG,BJ
광파오븐/식기세척기/전기레인지: 3yr: J(1대),L(2-3),O(4-9),R(10-29),U(30+) | 4yr: Y,AA,AD,AG,AJ | 5yr: AN,AP,AS,AV,AY | 6yr: BC,BE,BH,BK,BN
"""
from openpyxl.utils import column_index_from_string as _col

PRICE_COLS_MAP = {
    '냉장고': {
        '36': {'1': _col('I'), '2-3': _col('K'), '4-9': _col('N'), '10-29': _col('Q'), '30+': _col('T')},
        '48': {'1': _col('W'), '2-3': _col('Y'), '4-9': _col('AB'), '10-29': _col('AE'), '30+': _col('AH')},
        '60': {'1': _col('AK'), '2-3': _col('AM'), '4-9': _col('AP'), '10-29': _col('AS'), '30+': _col('AV')},
        '72': {'1': _col('AY'), '2-3': _col('BA'), '4-9': _col('BD'), '10-29': _col('BG'), '30+': _col('BJ')},
    },
    '광파오븐': {
        '36': {'1': _col('J'), '2-3': _col('L'), '4-9': _col('O'), '10-29': _col('R'), '30+': _col('U')},
        '48': {'1': _col('Y'), '2-3': _col('AA'), '4-9': _col('AD'), '10-29': _col('AG'), '30+': _col('AJ')},
        '60': {'1': _col('AN'), '2-3': _col('AP'), '4-9': _col('AS'), '10-29': _col('AV'), '30+': _col('AY')},
        '72': {'1': _col('BC'), '2-3': _col('BE'), '4-9': _col('BH'), '10-29': _col('BK'), '30+': _col('BN')},
    },
}
PRICE_COLS_MAP['식기세척기'] = PRICE_COLS_MAP['광파오븐']
PRICE_COLS_MAP['전기레인지'] = PRICE_COLS_MAP['광파오븐']

DATA_START = 9


def parse_type_c(ws, sheet_name):
    price_cols = PRICE_COLS_MAP.get(sheet_name, PRICE_COLS_MAP['냉장고'])
    products = []
    current_cat1 = ''
    current_cat2 = ''

    for row_idx in range(DATA_START, ws.max_row + 1):
        d_val = ws.cell(row=row_idx, column=_col('D')).value
        e_val = ws.cell(row=row_idx, column=_col('E')).value
        f_val = ws.cell(row=row_idx, column=_col('F')).value
        g_val = ws.cell(row=row_idx, column=_col('G')).value
        h_val = ws.cell(row=row_idx, column=_col('H')).value

        if d_val:
            current_cat1 = str(d_val).strip().replace('\n', ' ')
        if e_val:
            current_cat2 = str(e_val).strip().replace('\n', ' ')

        if not f_val:
            continue

        model_id = str(f_val).strip()
        service_type = str(g_val).strip() if g_val else ''
        visit_cycle = ''
        if h_val is not None:
            try:
                vc = int(h_val)
                visit_cycle = f"{vc}개월" if vc > 0 else "방문없음"
            except (ValueError, TypeError):
                visit_cycle = str(h_val).strip()

        prices = {}
        for period, tier_cols in price_cols.items():
            period_prices = {}
            for tier, col_idx in tier_cols.items():
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None and isinstance(val, (int, float)):
                    period_prices[tier] = int(val)
            if period_prices:
                prices[period] = period_prices

        if not prices:
            continue

        product_name = f"{current_cat1} {current_cat2}".strip()

        products.append({
            'model_id': model_id,
            'product_name': product_name,
            'service_type': service_type,
            'visit_cycle': visit_cycle,
            'prices': prices,
        })

    return products
